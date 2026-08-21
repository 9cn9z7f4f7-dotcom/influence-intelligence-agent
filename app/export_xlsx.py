from __future__ import annotations

from io import BytesIO
from datetime import date, datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1F2937')
HEADER_FONT = Font(color='FFFFFF', bold=True)
LINK_FONT = Font(color='008000', underline='single')


def _cell_value(value):
    """Convert arbitrary analysis values into values openpyxl can safely write."""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if hasattr(value, "value") and isinstance(getattr(value, "value", None), (str, int, float, bool)):
        return value.value
    if isinstance(value, (dict, list, tuple, set)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            return str(value)
    return str(value)


def _write_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center')
    for row in rows:
        ws.append([_cell_value(value) for value in row])
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, 1):
        width = max(12, min(55, max(len(str(header)), *(len(str(ws.cell(r, col_idx).value or '')) for r in range(2, ws.max_row + 1))) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    # URL columns become clickable.
    for col_idx, header in enumerate(headers, 1):
        if 'URL' in str(header).upper() or 'ССЫЛК' in str(header).upper():
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, col_idx)
                value = str(cell.value or '')
                if value.startswith(('http://', 'https://')):
                    cell.hyperlink = value
                    cell.font = LINK_FONT


def build_analysis_xlsx(result) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Обзор'
    summary = result.summary
    rows = [
        ['Бренд', result.brand.canonical_name],
        ['Создано', result.created_at],
        ['Материалов найдено', len(result.findings or [])],
        ['Подтверждённых интеграций', summary.confirmed_integrations],
        ['Авторов в выборке', summary.creators_used],
        ['Кандидатов с органическим интересом', summary.potential_creators_count],
        ['Площадки', ', '.join(p.value if hasattr(p, 'value') else str(p) for p in result.platforms)],
    ]
    for row in rows:
        ws.append([_cell_value(value) for value in row])
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 60
    for c in ws['A']:
        c.font = Font(bold=True)

    findings = []
    for f in result.findings or []:
        findings.append([
            f.get('entity_name'), f.get('entity_type'), f.get('platform'), f.get('content_title'),
            f.get('topic'), f.get('format'), f.get('classification'), f.get('source_url'),
        ])
    ws = wb.create_sheet('Что нашли')
    _write_table(ws, ['Автор / источник', 'Тип', 'Площадка', 'Материал', 'Тема', 'Формат', 'Классификация', 'URL источника'], findings)

    hunting = []
    for block in result.next_move or []:
        for c in block.get('candidates', []) or []:
            hunting.append([
                c.get('candidate'), c.get('platform'), c.get('followers'), c.get('median_views') or c.get('avg_views'),
                c.get('similarity_score'), c.get('reason') or c.get('why') or '', c.get('profile_url') or c.get('canonical_url'),
            ])
    ws = wb.create_sheet('Хант')
    _write_table(ws, ['Автор', 'Площадка', 'Подписчики', 'Просмотры', 'Match', 'Почему подходит', 'URL профиля'], hunting)

    segments = []
    for s in (result.white_space or {}).get('segments', []) or []:
        seg = s.get('segment') or {}
        segments.append([
            seg.get('label') or seg.get('topic'), s.get('creator_supply'), s.get('confirmed_integrations'),
            s.get('saturation_score'), s.get('opportunity_score'), bool(s.get('insufficient_data')),
        ])
    ws = wb.create_sheet('White Space')
    _write_table(ws, ['Сегмент', 'Авторов', 'Интеграций', 'Насыщенность', 'Возможность', 'Недостаточно данных'], segments)

    actions = []
    for item in (result.our_move or {}).get('opportunities', []) or []:
        actions.append([item.get('title') or item.get('action') or '', item.get('description') or item.get('reason') or ''])
    ws = wb.create_sheet('Что проверить')
    _write_table(ws, ['Действие', 'Почему'], actions)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
